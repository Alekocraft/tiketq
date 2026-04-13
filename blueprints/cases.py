from __future__ import annotations

import io
import os
import re
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

from flask import (
    Blueprint,
    current_app,
    abort,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from flask_login import current_user, login_required

from services.app_logging import log_event
from services.db import commit, execute, get_db, rollback, select_all, select_one
from services.ldap_auth import search_user
from services.mail import send_mail
from services.security import path_text, public_error_message, public_mail_message, secure_status_code, text_value
from services.roles import can_access_general_cases, can_ingest, can_resolve, can_triage, is_admin, normalize_role, normalize_roles, role_choices, role_label, team_aliases_for_roles, triage_targets_for_roles
from services.sla import compute_due_dates, get_priority_defaults, humanize_minutes, normalize_priority, priority_choices

cases_bp = Blueprint("cases", __name__, url_prefix="")

WAITING_USER_STATUS = "en espera de usuario"
OPEN_STATUSES = {"pendiente", "asignado", "reabierto", WAITING_USER_STATUS}
RESOLVED_STATUSES = {"resuelto"}
HIDDEN_STATUSES = {"cerrado"}
FINAL_STATUSES = RESOLVED_STATUSES | HIDDEN_STATUSES
PRIORITY_CHOICES = [choice["key"] for choice in priority_choices()]
_ATTACHMENT_NAME_RE = re.compile(r"[^A-Za-z0-9._-]+")


def _fetchall_dict(cur):
    rows = cur.fetchall()
    if not rows:
        return []
    cols = [c[0] for c in cur.description]
    return [dict(zip(cols, r)) for r in rows]


def _fetchone_dict(cur):
    row = cur.fetchone()
    if not row:
        return None
    cols = [c[0] for c in cur.description]
    return dict(zip(cols, row))


def _current_roles() -> list[str]:
    return normalize_roles(
        getattr(current_user, "roles", [])
        or ([getattr(current_user, "role", "")] if getattr(current_user, "role", "") else [])
    )


def _is_admin() -> bool:
    return is_admin(_current_roles())


def _is_manager() -> bool:
    return _is_admin()


def _can_triage() -> bool:
    return can_triage(_current_roles())


def _can_ingest() -> bool:
    return can_ingest(_current_roles())


def _can_resolve() -> bool:
    return can_resolve(_current_roles())


def _dedupe(items):
    result = []
    for item in items:
        if item and item not in result:
            result.append(item)
    return result


def _triage_team_choices() -> list[dict]:
    targets = triage_targets_for_roles(_current_roles())
    return [choice for choice in role_choices() if choice["key"] in targets]


def _priority_choices() -> list[dict]:
    return priority_choices()


def _fmt_dt(value):
    if not value:
        return "-"
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d %H:%M")
    return f"{value}"


def _safe_filename(name: str) -> str:
    cleaned = _ATTACHMENT_NAME_RE.sub("_", (name or "").strip()) or "adjunto"
    return cleaned[:180]




def _upload_root_for_case(case_id: str) -> Path:
    configured_root = (os.getenv("UPLOAD_ROOT") or "").strip()
    if configured_root:
        base_root = Path(configured_root)
        if not base_root.is_absolute():
            base_root = Path(current_app.root_path) / base_root
    else:
        base_root = Path(current_app.instance_path) / "uploads"
    upload_root = (base_root / case_id).resolve()
    upload_root.mkdir(parents=True, exist_ok=True)
    return upload_root


def _next_available_path(directory: Path, filename: str) -> Path:
    filepath = directory / filename
    counter = 1
    while filepath.exists():
        filepath = directory / f"{filepath.stem}_{counter}{filepath.suffix}"
        counter += 1
    return filepath


def _save_uploaded_files(case_id: str, files, update_id: Optional[int] = None):
    saved = []
    if not files:
        return saved

    upload_root = None
    for storage in files:
        if not storage:
            continue
        raw_name = (getattr(storage, "filename", "") or "").strip()
        if not raw_name:
            continue
        filename = _safe_filename(raw_name)
        if not filename:
            continue
        upload_root = upload_root or _upload_root_for_case(case_id)
        filepath = _next_available_path(upload_root, filename)
        storage.save(path_text(filepath))
        size_bytes = int(filepath.stat().st_size) if filepath.exists() else 0
        execute(
            """
            INSERT INTO dbo.case_attachments
                (case_id, update_id, filename, stored_path, content_type, size_bytes)
            VALUES (?,?,?,?,?,?)
            """,
            (
                case_id,
                update_id,
                filepath.name,
                path_text(filepath),
                (getattr(storage, "mimetype", None) or None),
                size_bytes,
            ),
        )
        saved.append(
            {
                "filename": filepath.name,
                "stored_path": path_text(filepath),
                "size_bytes": size_bytes,
                "update_id": update_id,
            }
        )
    return saved


def _decorate_case(row: Optional[dict]):
    if not row:
        return row
    priority = normalize_priority(row.get("priority") or "MEDIA")
    response_min, resolution_min = get_priority_defaults(priority)
    response_min = int(row.get("sla_response_min") or response_min)
    resolution_min = int(row.get("sla_resolution_min") or resolution_min)
    created_at = row.get("created_at") or datetime.now()
    response_due = row.get("response_due_at")
    resolution_due = row.get("resolution_due_at")
    if not response_due or not resolution_due:
        response_due, resolution_due = compute_due_dates(created_at, response_min, resolution_min)
    row["priority"] = priority
    row["priority_label"] = priority.title()
    row["priority_css"] = priority.lower()
    row["sla_response_min"] = response_min
    row["sla_resolution_min"] = resolution_min
    row["sla_response_text"] = humanize_minutes(response_min)
    row["sla_resolution_text"] = humanize_minutes(resolution_min)
    row["response_due_at"] = response_due
    row["resolution_due_at"] = resolution_due
    row["response_due_at_display"] = _fmt_dt(response_due)
    row["resolution_due_at_display"] = _fmt_dt(resolution_due)
    row["sla_window_text"] = f"{row['sla_response_text']} / {row['sla_resolution_text']}"
    row["assigned_team"] = normalize_role(row.get("assigned_team") or "")
    row["assigned_team_label"] = role_label(row.get("assigned_team") or "") if row.get("assigned_team") else "-"
    return row


def _work_queue_aliases() -> list[str]:
    return _dedupe(team_aliases_for_roles(_current_roles()))


def _visibility_condition(alias: str = "c", scope: str = "all"):
    if _is_admin():
        if scope == "mine":
            return f"LOWER(ISNULL({alias}.assigned_to, '')) = LOWER(?)", [current_user.username]
        return "", []

    team_aliases = _work_queue_aliases()
    user_clause = f"LOWER(ISNULL({alias}.assigned_to, '')) = LOWER(?)"
    if not team_aliases:
        return user_clause, [current_user.username]

    expr = f"LOWER(REPLACE(REPLACE(ISNULL({alias}.assigned_team, ''), ' ', '_'), '-', '_'))"
    placeholders = ", ".join(["?"] * len(team_aliases))
    team_clause = f"{expr} IN ({placeholders})"
    return f"({team_clause} OR {user_clause})", [*team_aliases, current_user.username]


def _notif_count() -> int:
    row = select_one(
        "SELECT COUNT(*) AS c FROM dbo.notifications WHERE user_id=? AND is_read=0",
        (current_user.username,),
    )
    return int(row["c"]) if row and row.get("c") is not None else 0


def _notification_visual(notification_type: str | None) -> dict:
    normalized = normalize_role(notification_type or "")
    tokens = set((normalized or "info").split("_"))

    if {"security", "seguridad", "auth", "acceso"} & tokens:
        return {"label": "Seguridad", "tone": "security", "icon": "🛡️"}
    if {"warning", "alerta", "critico", "critical", "sla", "urgent"} & tokens:
        return {"label": "Alerta", "tone": "warning", "icon": "⚠️"}
    if {"success", "resolved", "resuelto", "ok", "done", "close", "cerrado"} & tokens:
        return {"label": "Completado", "tone": "success", "icon": "✓"}
    if {"case", "ticket", "caso", "update", "asignado", "reabierto"} & tokens:
        return {"label": "Caso", "tone": "case", "icon": "📌"}
    return {"label": "Información", "tone": "info", "icon": "ℹ️"}


def _notification_summary(rows: list[dict]) -> dict:
    total = len(rows)
    unread = sum(1 for row in rows if not bool(row.get("is_read")))
    read = total - unread
    grouped: dict[str, dict] = {}
    for row in rows:
        visual = row.get("visual") or _notification_visual(row.get("type"))
        key = visual["label"]
        if key not in grouped:
            grouped[key] = {"label": visual["label"], "tone": visual["tone"], "count": 0}
        grouped[key]["count"] += 1

    highlights = sorted(grouped.values(), key=lambda item: (-item["count"], item["label"]))[:4]
    return {"total": total, "unread": unread, "read": read, "highlights": highlights}


def _role_scope_labels() -> list[str]:
    roles = _current_roles()
    if _is_admin():
        return ["Administrador"]
    return [role_label(r) for r in roles] or ["Sin rol asignado"]


def _dashboard_stats():
    db = get_db()
    visibility_sql, visibility_params = _visibility_condition("c", scope="all")
    mine_sql, mine_params = _visibility_condition("c", scope="mine")

    filters = ["LOWER(ISNULL(c.status, '')) <> 'cerrado'"]
    params = list(visibility_params)
    if visibility_sql:
        filters.insert(0, visibility_sql)
    where_clause = "WHERE " + " AND ".join(filters)
    cur = db.cursor()
    cur.execute(
        f"""
        SELECT
            COUNT(*) AS total_cases,
            SUM(CASE WHEN LOWER(c.status) IN ('pendiente', 'asignado', 'reabierto', 'en espera de usuario') THEN 1 ELSE 0 END) AS open_cases,
            SUM(CASE WHEN LOWER(c.status) = 'resuelto' THEN 1 ELSE 0 END) AS resolved_cases,
            SUM(CASE WHEN LOWER(c.status) = 'reabierto' THEN 1 ELSE 0 END) AS reopened_cases,
            SUM(CASE WHEN UPPER(c.priority) IN ('ALTA', 'P1') THEN 1 ELSE 0 END) AS high_priority
        FROM dbo.cases c
        {where_clause}
        """,
        params,
    )
    stats = _fetchone_dict(cur) or {}

    mine_filters = ["LOWER(ISNULL(c.status, '')) <> 'cerrado'"]
    mine_bindings = []
    if mine_sql:
        mine_filters.insert(0, mine_sql)
        mine_bindings.extend(mine_params)
    else:
        mine_filters.insert(0, "LOWER(ISNULL(c.assigned_to, '')) = LOWER(?)")
        mine_bindings.append(current_user.username)
    mine_where = "WHERE " + " AND ".join(mine_filters)
    mine_row = select_one(
        f"SELECT COUNT(*) AS total_cases FROM dbo.cases c {mine_where}",
        tuple(mine_bindings),
    ) or {}

    cur_recent = db.cursor()
    cur_recent.execute(
        f"""
        SELECT TOP 8
            c.id, c.subject, c.status, c.priority, c.assigned_team,
            c.assigned_to, c.requester_email, c.created_at, c.updated_at
        FROM dbo.cases c
        {where_clause}
        ORDER BY c.created_at DESC
        """,
        params,
    )
    recent_cases = _fetchall_dict(cur_recent)

    cur_ingest = db.cursor()
    cur_ingest.execute(
        """
        SELECT TOP 1 processed_at, status, case_id, email_message_id
        FROM dbo.email_ingest_log
        ORDER BY processed_at DESC
        """
    )
    last_ingest = _fetchone_dict(cur_ingest)

    return {
        "total_cases": int(stats.get("total_cases") or 0),
        "open_cases": int(stats.get("open_cases") or 0),
        "resolved_cases": int(stats.get("resolved_cases") or 0),
        "reopened_cases": int(stats.get("reopened_cases") or 0),
        "high_priority": int(stats.get("high_priority") or 0),
        "my_cases": int(mine_row.get("total_cases") or 0),
        "recent_cases": recent_cases,
        "last_ingest": last_ingest,
        "notif_count": _notif_count(),
        "scope_labels": _role_scope_labels(),
    }


def _list_cases(scope: str = "all"):
    db = get_db()
    status = (request.args.get("status") or "").strip()
    q = (request.args.get("q") or "").strip()
    assigned_team = normalize_role(request.args.get("assigned_team") or "")

    try:
        limit = int(request.args.get("limit", 50))
    except ValueError:
        limit = 50
    limit = max(1, min(limit, 200))

    where = []
    params = []

    visibility_sql, visibility_params = _visibility_condition("c", scope="mine" if scope == "mine" else "all")
    where.append("LOWER(ISNULL(c.status, '')) <> 'cerrado'")

    if visibility_sql:
        where.append(visibility_sql)
        params.extend(visibility_params)

    if status and status.strip().lower() != "cerrado":
        where.append("LOWER(c.status) = LOWER(?)")
        params.append(status)

    if q:
        where.append("(c.id LIKE ? OR c.subject LIKE ? OR c.requester_email LIKE ?)")
        params.extend([f"%{q}%"] * 3)

    if assigned_team and _is_admin() and scope != "mine":
        where.append("LOWER(REPLACE(REPLACE(ISNULL(c.assigned_team, ''), ' ', '_'), '-', '_')) = ?")
        params.append(assigned_team)

    sql = """
        SELECT
            c.id, c.subject, c.description, c.status, c.priority, c.category, c.subcategory,
            c.assigned_team, c.assigned_to,
            u.display_name AS assigned_user_name,
            c.requester_name, c.requester_email,
            c.sla_response_min, c.sla_resolution_min, c.response_due_at, c.resolution_due_at,
            c.created_at, c.updated_at, c.first_response_at, c.resolved_at, c.closed_at
        FROM dbo.cases c
        LEFT JOIN dbo.users u ON u.id = c.assigned_to
    """

    if where:
        sql += " WHERE " + " AND ".join(where)

    sql += " ORDER BY c.created_at DESC OFFSET 0 ROWS FETCH NEXT ? ROWS ONLY"
    params_with_limit = params + [limit]

    cur = db.cursor()
    cur.execute(sql, params_with_limit)
    cases = [_decorate_case(row) for row in _fetchall_dict(cur)]

    summary_clause = " WHERE " + " AND ".join(where) if where else ""
    cur_summary = db.cursor()
    cur_summary.execute(
        f"""
        SELECT
            COUNT(*) AS total_cases,
            SUM(CASE WHEN LOWER(c.status) IN ('pendiente', 'asignado', 'reabierto', 'en espera de usuario') THEN 1 ELSE 0 END) AS open_cases,
            SUM(CASE WHEN LOWER(c.status) = 'resuelto' THEN 1 ELSE 0 END) AS resolved_cases,
            SUM(CASE WHEN LOWER(c.status) = 'reabierto' THEN 1 ELSE 0 END) AS reopened_cases
        FROM dbo.cases c
        {summary_clause}
        """,
        params,
    )
    summary = _fetchone_dict(cur_summary) or {}

    return render_template(
        "cases/list.html",
        title="Mis tickets" if scope == "mine" else "Casos",
        cases=cases,
        notif_count=_notif_count(),
        status=status,
        q=q,
        limit=limit,
        summary=summary,
        scope=scope,
        assigned_team=assigned_team,
        role_choices=role_choices(),
        scope_labels=_role_scope_labels(),
        can_triage=_can_triage(),
    )


def _get_case(case_id: str):
    visibility_sql, visibility_params = _visibility_condition("c", scope="all")
    sql = """
        SELECT
            c.id, c.subject, c.description, c.status, c.priority, c.category, c.subcategory,
            c.assigned_team, c.assigned_to, c.requester_name, c.requester_email,
            c.sla_response_min, c.sla_resolution_min, c.response_due_at, c.resolution_due_at,
            c.created_at, c.updated_at, c.first_response_at, c.resolved_at, c.closed_at,
            u.display_name AS assigned_user_name
        FROM dbo.cases c
        LEFT JOIN dbo.users u ON u.id = c.assigned_to
        WHERE c.id = ?
    """
    params = [case_id]
    if visibility_sql:
        sql += f" AND {visibility_sql}"
        params.extend(visibility_params)
    return _decorate_case(select_one(sql, tuple(params)))


def _record_case_update(
    case_id: str,
    message: str,
    is_solution: bool = False,
    *,
    author_id: str | None = None,
    author_name: str | None = None,
    author_email: str | None = None,
):
    cur = execute(
        """
        INSERT INTO dbo.case_updates(case_id, author_id, author_name, author_email, message, is_solution, created_at)
        OUTPUT INSERTED.id
        VALUES (?, ?, ?, ?, ?, ?, SYSDATETIME())
        """,
        (
            case_id,
            current_user.username if author_id is None else author_id,
            current_user.display_name if author_name is None else author_name,
            current_user.email if author_email is None else author_email,
            message,
            1 if is_solution else 0,
        ),
    )
    row = cur.fetchone()
    if not row:
        return None
    return int(row[0])


@cases_bp.route("/dashboard")
@login_required
def dashboard():
    data = _dashboard_stats()
    data["summary"] = {
        "total_cases": data.get("total_cases", 0),
        "open_cases": data.get("open_cases", 0),
        "resolved_cases": data.get("resolved_cases", 0),
        "reopened_cases": data.get("reopened_cases", 0),
        "high_priority": data.get("high_priority", 0),
        "my_cases": data.get("my_cases", 0),
    }
    data["latest_ingest"] = data.get("last_ingest")
    data["active_roles"] = _role_scope_labels()
    return render_template("dashboard.html", title="Inicio", **data)


@cases_bp.route("/cases")
@login_required
def list_cases():
    if not can_access_general_cases(_current_roles()):
        flash("Tu perfil solo tiene acceso a Mis tickets.", "error")
        return redirect(url_for("cases.my_tickets"))
    return _list_cases("all")


@cases_bp.route("/mis-tickets")
@login_required
def my_tickets():
    return _list_cases("mine")


@cases_bp.route("/cases/categorizados")
@login_required
def categorized_cases():
    return redirect(url_for("cases.my_tickets"))

@cases_bp.route("/cases/<case_id>")
@login_required
def case_detail(case_id):
    case_row = _get_case(case_id)
    if not case_row:
        abort(404)

    attachments = select_all(
        "SELECT id, update_id, filename, stored_path, size_bytes FROM dbo.case_attachments WHERE case_id = ? ORDER BY id DESC",
        (case_id,),
    )
    updates = select_all(
        """
        SELECT id, author_name, author_email, message, is_solution, created_at
        FROM dbo.case_updates
        WHERE case_id = ?
        ORDER BY created_at DESC
        """,
        (case_id,),
    )
    attachments_by_update = {}
    for attachment in attachments:
        update_id = attachment.get("update_id")
        if update_id is None:
            continue
        attachments_by_update.setdefault(int(update_id), []).append(attachment)

    for item in updates:
        item["attachments"] = attachments_by_update.get(int(item.get("id") or 0), [])

    latest_survey = select_one(
        """
        SELECT TOP 1 resolved_at_snapshot, sent_at, rating, reason, completed_at
        FROM dbo.case_surveys
        WHERE case_id = ?
        ORDER BY created_at DESC, id DESC
        """,
        (case_id,),
    )
    if latest_survey:
        latest_survey["sent_at_display"] = _fmt_dt(latest_survey.get("sent_at"))
        latest_survey["completed_at_display"] = _fmt_dt(latest_survey.get("completed_at"))

    return render_template(
        "cases/detail.html",
        title=f"Caso {case_id}",
        c=case_row,
        atts=attachments,
        updates=updates,
        latest_survey=latest_survey,
        notif_count=_notif_count(),
        can_reopen=(_can_resolve() and text_value(case_row.get("status")).lower() in FINAL_STATUSES),
        can_resolve=(_can_resolve() and text_value(case_row.get("status")).lower() not in FINAL_STATUSES),
        can_wait_user=(_can_resolve() and text_value(case_row.get("status")).lower() not in FINAL_STATUSES),
        can_close=(_can_resolve() and text_value(case_row.get("status")).lower() in RESOLVED_STATUSES),
        can_triage=_can_triage(),
        triage_teams=_triage_team_choices(),
        priority_choices=_priority_choices(),
    )




def _resolve_attachment_path(stored_path: str) -> Path | None:
    raw = (stored_path or "").strip()
    if not raw:
        return None

    path = Path(raw)
    candidates = []
    if path.is_absolute():
        candidates.append(path)
    else:
        candidates.extend([
            Path(current_app.root_path) / path,
            Path(current_app.instance_path) / path,
            Path.cwd() / path,
        ])
        uploads_name = path.parts[0] if path.parts else "uploads"
        if uploads_name:
            remainder = Path(*path.parts[1:]) if len(path.parts) > 1 else Path()
            candidates.append(Path(current_app.instance_path) / uploads_name / remainder)

    for candidate in candidates:
        try:
            resolved = candidate.resolve()
        except Exception:
            resolved = candidate
        if resolved.exists() and resolved.is_file():
            return resolved
    return None


@cases_bp.route("/cases/<case_id>/attachments/id/<int:attachment_id>")
@login_required
def download_attachment_by_id(case_id, attachment_id):
    case_row = _get_case(case_id)
    if not case_row:
        abort(404)
    row = select_one(
        "SELECT TOP 1 id, filename, stored_path FROM dbo.case_attachments WHERE case_id = ? AND id = ?",
        (case_id, attachment_id),
    )
    if not row:
        abort(404)
    resolved_path = _resolve_attachment_path(row.get("stored_path") or "")
    if not resolved_path:
        abort(404)
    return send_file(path_text(resolved_path), as_attachment=True, download_name=row.get("filename") or resolved_path.name)


@cases_bp.route("/cases/<case_id>/attachments/<path:filename>")
@login_required
def download_attachment(case_id, filename):
    case_row = _get_case(case_id)
    if not case_row:
        abort(404)
    row = select_one(
        "SELECT TOP 1 id, filename, stored_path FROM dbo.case_attachments WHERE case_id = ? AND filename = ? ORDER BY id DESC",
        (case_id, filename),
    )
    if not row:
        abort(404)
    resolved_path = _resolve_attachment_path(row.get("stored_path") or "")
    if not resolved_path:
        abort(404)
    return send_file(path_text(resolved_path), as_attachment=True, download_name=row.get("filename") or filename)




def _redirect_case_detail(case_id: str):
    return redirect(url_for("cases.case_detail", case_id=case_id))


def _mail_attachments_from_saved_files(saved_files) -> list[dict]:
    attachments = []
    for item in saved_files or []:
        stored_path = (item or {}).get("stored_path") or ""
        resolved_path = _resolve_attachment_path(stored_path)
        if not resolved_path:
            continue
        attachments.append(
            {
                "path": path_text(resolved_path),
                "filename": (item or {}).get("filename") or resolved_path.name,
                "content_type": (item or {}).get("content_type") or None,
            }
        )
    return attachments


def _notify_requester_case_resolved(case_row: dict, note: str, saved_files=None) -> tuple[bool, str]:
    recipient = (case_row.get("requester_email") or "").strip()
    if not recipient:
        return False, "El caso quedó resuelto, pero el solicitante no tiene correo registrado."

    requester_name = (case_row.get("requester_name") or "").strip()
    greeting = f"Hola {requester_name}," if requester_name else "Hola,"
    subject = (case_row.get("subject") or "Sin asunto").strip()
    analyst_name = (getattr(current_user, "display_name", None) or getattr(current_user, "username", None) or "Mesa de ayuda").strip()

    body = (
        f"{greeting}\n\n"
        f"Tu caso {case_row['id']} fue marcado como RESUELTO.\n"
        f"Asunto: {subject}\n\n"
        "Detalle de la solución o gestión realizada:\n"
        f"{note}\n\n"
        f"Gestionado por: {analyst_name}\n\n"
        "Si necesitas más ayuda, responde a este correo y el caso podrá ser revisado nuevamente.\n\n"
        "Gracias,\n"
        "Mesa de ayuda Qualitas"
    )

    try:
        ok, status = send_mail(
            to_addr=recipient,
            subject=f"[{case_row['id']}] Caso resuelto - Soporte TI",
            body=body,
            attachments=_mail_attachments_from_saved_files(saved_files),
        )
    except Exception as exc:
        return False, public_mail_message("MAIL_SEND_FAILED")

    return bool(ok), "" if ok else public_mail_message(status)


def _handle_status_change(case_id: str, case_row: dict, action: str, note: str, upload_files):
    if action not in {"resolve", "reopen", "close", "wait_user"}:
        flash("Acción no reconocida.", "warning")
        return _redirect_case_detail(case_id)

    if not note:
        flash("La observación es obligatoria para actualizar el estado del caso.", "warning")
        return _redirect_case_detail(case_id)

    try:
        notify_result = None
        if action == "resolve":
            execute(
                "UPDATE dbo.cases SET status='RESUELTO', resolved_at=SYSDATETIME(), updated_at=SYSDATETIME() WHERE id = ?",
                (case_id,),
            )
            update_id = _record_case_update(case_id, note, is_solution=True)
            saved_files = _save_uploaded_files(case_id, upload_files, update_id=update_id)
            notify_result = (case_row, note, saved_files)
            flash(f"El caso {case_id} quedó en estado RESUELTO.", "success")
        elif action == "wait_user":
            execute(
                "UPDATE dbo.cases SET status='EN ESPERA DE USUARIO', updated_at=SYSDATETIME() WHERE id = ?",
                (case_id,),
            )
            update_id = _record_case_update(case_id, note)
            _save_uploaded_files(case_id, upload_files, update_id=update_id)
            flash(f"El caso {case_id} quedó en estado EN ESPERA DE USUARIO.", "success")
        elif action == "close":
            execute(
                "UPDATE dbo.cases SET status='CERRADO', closed_at=SYSDATETIME(), updated_at=SYSDATETIME() WHERE id = ?",
                (case_id,),
            )
            update_id = _record_case_update(case_id, note)
            _save_uploaded_files(case_id, upload_files, update_id=update_id)
            flash(f"El caso {case_id} quedó en estado CERRADO.", "success")
        else:
            response_min, resolution_min = get_priority_defaults(case_row.get("priority") or "MEDIA")
            response_due, resolution_due = compute_due_dates(datetime.now(), response_min, resolution_min)
            execute(
                "UPDATE dbo.cases SET status='REABIERTO', resolved_at=NULL, closed_at=NULL, response_due_at=?, resolution_due_at=?, updated_at=SYSDATETIME() WHERE id = ?",
                (response_due, resolution_due, case_id),
            )
            update_id = _record_case_update(case_id, note)
            _save_uploaded_files(case_id, upload_files, update_id=update_id)
            flash(f"El caso {case_id} fue reabierto.", "success")
        commit()
        log_event(
            "CASE",
            "INFO",
            "CASE_STATUS_CHANGED",
            detail=f"accion {action}",
            source="cases._handle_status_change",
            user_id=current_user.username,
            case_id=case_id,
            status="OK",
            metadata={"attachments": len(saved_files) if action == "resolve" else 0},
        )
        if notify_result:
            ok, detail = _notify_requester_case_resolved(*notify_result)
            if ok:
                flash("Se notificó al solicitante por correo que el caso quedó resuelto.", "success")
            else:
                flash(f"El caso quedó resuelto, pero no fue posible enviar la notificación: {public_mail_message()}", "warning")
    except Exception as exc:
        rollback()
        log_event(
            "CASE",
            "ERROR",
            "CASE_STATUS_CHANGE_FAILED",
            detail="UNEXPECTED_ERROR",
            source="cases._handle_status_change",
            user_id=current_user.username,
            case_id=case_id,
            status="FAILED",
        )
        flash(public_error_message("No fue posible actualizar el caso."), "error")
    return _redirect_case_detail(case_id)


def _handle_triage(case_id: str, assigned_team: str, priority: str, note: str, upload_files):
    allowed_targets = triage_targets_for_roles(_current_roles())
    if assigned_team not in allowed_targets:
        flash("Debes seleccionar un destino permitido para tu rol.", "warning")
        return _redirect_case_detail(case_id)
    if priority not in PRIORITY_CHOICES:
        flash("Debes seleccionar una prioridad válida.", "warning")
        return _redirect_case_detail(case_id)

    try:
        response_min, resolution_min = get_priority_defaults(priority)
        response_due, resolution_due = compute_due_dates(datetime.now(), response_min, resolution_min)
        execute(
            """
            UPDATE dbo.cases
            SET assigned_team=?,
                assigned_to=?,
                priority=?,
                category=?,
                sla_response_min=?,
                sla_resolution_min=?,
                response_due_at=?,
                resolution_due_at=?,
                status=CASE WHEN LOWER(status)='resuelto' THEN status ELSE 'ASIGNADO' END,
                updated_at=SYSDATETIME()
            WHERE id=?
            """,
            (assigned_team, None, priority, assigned_team, int(response_min), int(resolution_min), response_due, resolution_due, case_id),
        )
        message = f"Caso asignado a la cola de {role_label(assigned_team)} con prioridad {priority}."
        if note:
            message += f" Nota: {note}"
        update_id = _record_case_update(case_id, message)
        _save_uploaded_files(case_id, upload_files, update_id=update_id)
        commit()
        flash("Caso asignado correctamente.", "success")
    except Exception as exc:
        rollback()
        flash(public_error_message("No fue posible asignar el caso."), "error")
    return _redirect_case_detail(case_id)


@cases_bp.route("/cases/<case_id>/work", methods=["POST"])
@login_required
def manage_case(case_id):
    case_row = _get_case(case_id)
    if not case_row:
        abort(404)

    operation = (request.form.get("operation") or request.form.get("action") or "").strip().lower()
    note = (request.form.get("note") or "").strip()
    upload_files = request.files.getlist("attachments")

    if operation == "triage":
        if not _can_triage():
            flash("No tienes permisos para reasignar o escalar este caso.", "error")
            return _redirect_case_detail(case_id)
        assigned_team = normalize_role(request.form.get("assigned_team") or "")
        priority = normalize_priority(request.form.get("priority") or "")
        return _handle_triage(case_id, assigned_team, priority, note, upload_files)

    if operation in {"resolve", "reopen", "close", "wait_user"}:
        if not _can_resolve():
            flash("No tienes permisos para actualizar el estado de este caso.", "error")
            return _redirect_case_detail(case_id)
        return _handle_status_change(case_id, case_row, operation, note, upload_files)

    flash("Debes seleccionar una acción válida.", "warning")
    return _redirect_case_detail(case_id)

@cases_bp.route("/cases/<case_id>/status", methods=["POST"])
@login_required
def change_case_status(case_id):
    case_row = _get_case(case_id)
    if not case_row:
        abort(404)
    if not _can_resolve():
        flash("No tienes permisos para actualizar el estado de este caso.", "error")
        return redirect(url_for("cases.case_detail", case_id=case_id))

    action = (request.form.get("action") or "").strip().lower()
    note = (request.form.get("note") or "").strip()
    upload_files = request.files.getlist("attachments")

    if action not in {"resolve", "reopen", "close", "wait_user"}:
        flash("Acción no reconocida.", "warning")
        return redirect(url_for("cases.case_detail", case_id=case_id))

    if not note:
        flash("La nota es obligatoria para actualizar el estado del caso.", "warning")
        return redirect(url_for("cases.case_detail", case_id=case_id))

    try:
        notify_result = None
        if action == "resolve":
            execute(
                "UPDATE dbo.cases SET status='RESUELTO', resolved_at=SYSDATETIME(), updated_at=SYSDATETIME() WHERE id = ?",
                (case_id,),
            )
            update_id = _record_case_update(case_id, note, is_solution=True)
            _save_uploaded_files(case_id, upload_files, update_id=update_id)
            notify_result = (case_row, note)
            flash(f"El caso {case_id} quedó en estado RESUELTO.", "success")
        elif action == "wait_user":
            execute(
                "UPDATE dbo.cases SET status='EN ESPERA DE USUARIO', updated_at=SYSDATETIME() WHERE id = ?",
                (case_id,),
            )
            update_id = _record_case_update(case_id, note)
            _save_uploaded_files(case_id, upload_files, update_id=update_id)
            flash(f"El caso {case_id} quedó en estado EN ESPERA DE USUARIO.", "success")
        elif action == "close":
            execute(
                "UPDATE dbo.cases SET status='CERRADO', closed_at=SYSDATETIME(), updated_at=SYSDATETIME() WHERE id = ?",
                (case_id,),
            )
            update_id = _record_case_update(case_id, note)
            _save_uploaded_files(case_id, upload_files, update_id=update_id)
            flash(f"El caso {case_id} quedó en estado CERRADO.", "success")
        else:
            response_min, resolution_min = get_priority_defaults(case_row.get("priority") or "MEDIA")
            response_due, resolution_due = compute_due_dates(datetime.now(), response_min, resolution_min)
            execute(
                "UPDATE dbo.cases SET status='REABIERTO', resolved_at=NULL, closed_at=NULL, response_due_at=?, resolution_due_at=?, updated_at=SYSDATETIME() WHERE id = ?",
                (response_due, resolution_due, case_id),
            )
            update_id = _record_case_update(case_id, note)
            _save_uploaded_files(case_id, upload_files, update_id=update_id)
            flash(f"El caso {case_id} fue reabierto.", "success")
        commit()
    except Exception as exc:
        rollback()
        flash(public_error_message("No fue posible actualizar el caso."), "error")

    return redirect(url_for("cases.case_detail", case_id=case_id))


@cases_bp.route("/cases/<case_id>/triage", methods=["POST"])
@login_required
def triage_case(case_id):
    if not _can_triage():
        flash("No tienes permisos para categorizar casos.", "error")
        return redirect(url_for("cases.case_detail", case_id=case_id))

    case_row = _get_case(case_id)
    if not case_row:
        abort(404)

    assigned_team = normalize_role(request.form.get("assigned_team") or "")
    priority = normalize_priority(request.form.get("priority") or "")
    note = (request.form.get("note") or "").strip()
    upload_files = request.files.getlist("attachments")

    allowed_targets = triage_targets_for_roles(_current_roles())
    if assigned_team not in allowed_targets:
        flash("Debes seleccionar un destino permitido para tu rol.", "warning")
        return redirect(url_for("cases.case_detail", case_id=case_id))
    if priority not in PRIORITY_CHOICES:
        flash("Debes seleccionar una prioridad válida.", "warning")
        return redirect(url_for("cases.case_detail", case_id=case_id))

    try:
        response_min, resolution_min = get_priority_defaults(priority)
        response_due, resolution_due = compute_due_dates(datetime.now(), response_min, resolution_min)
        execute(
            """
            UPDATE dbo.cases
            SET assigned_team=?,
                assigned_to=?,
                priority=?,
                category=?,
                sla_response_min=?,
                sla_resolution_min=?,
                response_due_at=?,
                resolution_due_at=?,
                status=CASE WHEN LOWER(status)='resuelto' THEN status ELSE 'ASIGNADO' END,
                updated_at=SYSDATETIME()
            WHERE id=?
            """,
            (assigned_team, None, priority, assigned_team, int(response_min), int(resolution_min), response_due, resolution_due, case_id),
        )
        message = f"Caso asignado a la cola de {role_label(assigned_team)} con prioridad {priority}."
        if note:
            message += f" Nota: {note}"
        update_id = _record_case_update(case_id, message)
        _save_uploaded_files(case_id, upload_files, update_id=update_id)
        commit()
        flash("Caso asignado correctamente.", "success")
    except Exception as exc:
        rollback()
        flash(public_error_message("No fue posible asignar el caso."), "error")

    return redirect(url_for("cases.case_detail", case_id=case_id))


@cases_bp.route("/actions/ingest-emails", methods=["POST"])
@login_required
def ingest_emails_action():
    if not _can_ingest():
        message = "Solo Administrador y Gestor TI pueden ejecutar la ingesta de correos."
        flash(message, "error")
        wants_json = (
            request.headers.get("X-Requested-With") == "XMLHttpRequest"
            or "application/json" in (request.headers.get("Accept") or "")
        )
        if wants_json:
            return jsonify(ok=False, error=message), 403
        return redirect(request.referrer or url_for("cases.dashboard"))

    wants_json = (
        request.headers.get("X-Requested-With") == "XMLHttpRequest"
        or "application/json" in (request.headers.get("Accept") or "")
    )

    try:
        from services.email_ingest import ingest_unseen
        result = ingest_unseen()
        created = int((result or {}).get("created") or 0)
        linked = int((result or {}).get("linked") or 0)

        if created > 0 or linked > 0:
            parts = []
            if created > 0:
                parts.append(f"se crearon {created} caso(s)")
            if linked > 0:
                parts.append(f"se asociaron {linked} correo(s) a casos existentes")
            flash(f"Ingesta completada: {' y '.join(parts)}.", "success")
        else:
            flash("Ingesta completada. No se encontraron correos nuevos para procesar.", "info")

        if wants_json:
            return jsonify(ok=True, result=result)
    except Exception as exc:
        rollback()
        message = public_error_message("La ingesta falló. Revisa la configuración y vuelve a intentarlo.")
        flash(message, "error")
        if wants_json:
            return jsonify(ok=False, error=public_error_message()), 500

    return redirect(request.referrer or url_for("cases.dashboard"))


@cases_bp.route("/notifications")
@login_required
def notifications():
    try:
        limit = int(request.args.get("limit", 50))
    except ValueError:
        limit = 50
    limit = max(1, min(limit, 200))

    notifications_rows = select_all(
        """
        SELECT id, type, title, body, case_id, is_read, created_at
        FROM dbo.notifications
        WHERE user_id=?
        ORDER BY created_at DESC
        OFFSET 0 ROWS FETCH NEXT ? ROWS ONLY
        """,
        (current_user.username, limit),
    )

    for row in notifications_rows:
        visual = _notification_visual(row.get("type"))
        row["visual"] = visual
        row["created_at_text"] = _fmt_dt(row.get("created_at"))

    summary = _notification_summary(notifications_rows)

    return render_template(
        "cases/notifications.html",
        title="Notificaciones",
        notifications=notifications_rows,
        notif_count=_notif_count(),
        notification_summary=summary,
        limit=limit,
    )


@cases_bp.route("/notifications/mark-read", methods=["POST"])
@login_required
def notifications_mark_read():
    cur = execute(
        "UPDATE dbo.notifications SET is_read=1, read_at=SYSDATETIME() WHERE user_id=? AND is_read=0",
        (current_user.username,),
    )
    affected = int(getattr(cur, "rowcount", 0) or 0)
    commit()
    log_event(
        "NOTIFICATION",
        "INFO",
        "NOTIFICATIONS_MARKED_READ",
        detail="notificaciones actualizadas",
        source="cases.notifications_mark_read",
        user_id=current_user.username,
        status="OK",
        metadata={"affected": affected},
    )
    return jsonify(ok=True, affected=affected)


@cases_bp.route("/admin/users/legacy", methods=["GET", "POST"])
@login_required
def admin_users():
    if not _is_admin():
        flash("Solo el perfil administrador puede gestionar usuarios.", "error")
        return redirect(url_for("cases.dashboard"))

    edit_username = (request.values.get("edit") or request.values.get("username") or request.values.get("user_id") or "").strip()
    if edit_username:
        return redirect(url_for("users.edit_user", user_id=edit_username))
    return redirect(url_for("users.list_users"))


@cases_bp.route("/encuesta/<token>", methods=["GET", "POST"])
def case_survey(token):
    survey = select_one(
        """
        SELECT TOP 1
            s.id, s.case_id, s.token, s.recipient_email, s.sent_at, s.rating, s.reason,
            s.completed_at, s.created_at,
            c.subject, c.requester_name
        FROM dbo.case_surveys s
        INNER JOIN dbo.cases c ON c.id = s.case_id
        WHERE s.token = ?
        """,
        (token,),
    )
    if not survey:
        abort(404)

    error_message = None
    success_message = None

    if request.method == "POST":
        if survey.get("completed_at"):
            success_message = "Esta encuesta ya fue respondida. Gracias por tu tiempo."
        else:
            rating_raw = (request.form.get("rating") or "").strip()
            reason = (request.form.get("reason") or "").strip()
            try:
                rating = int(rating_raw)
            except (TypeError, ValueError):
                rating = 0

            if rating not in {1, 2, 3, 4, 5}:
                error_message = "Selecciona una calificación válida entre 1 y 5."
            else:
                try:
                    execute(
                        """
                        UPDATE dbo.case_surveys
                        SET rating = ?, reason = ?, completed_at = SYSDATETIME(), updated_at = SYSDATETIME()
                        WHERE id = ?
                        """,
                        (rating, reason or None, survey["id"]),
                    )
                    audit_message = f"Encuesta de satisfacción respondida con calificación {rating}/5."
                    if reason:
                        audit_message += f" Motivo registrado: {reason}"
                    execute(
                        """
                        INSERT INTO dbo.case_updates(case_id, author_id, author_name, author_email, message, is_solution, created_at)
                        VALUES (?, NULL, ?, ?, ?, 0, SYSDATETIME())
                        """,
                        (
                            survey["case_id"],
                            survey.get("requester_name") or "Usuario",
                            survey.get("recipient_email") or None,
                            audit_message,
                        ),
                    )
                    commit()
                    log_event(
                        "AUDIT",
                        "INFO",
                        "SURVEY_COMPLETED",
                        detail=f"calificacion {rating}",
                        source="cases.survey",
                        case_id=survey["case_id"],
                        status="OK",
                        metadata={"rating": rating, "has_reason": bool(reason)},
                    )
                    success_message = "Gracias por responder la encuesta."
                    survey = select_one(
                        """
                        SELECT TOP 1
                            s.id, s.case_id, s.token, s.recipient_email, s.sent_at, s.rating, s.reason,
                            s.completed_at, s.created_at,
                            c.subject, c.requester_name
                        FROM dbo.case_surveys s
                        INNER JOIN dbo.cases c ON c.id = s.case_id
                        WHERE s.token = ?
                        """,
                        (token,),
                    )
                except Exception as exc:
                    rollback()
                    error_message = public_error_message("No fue posible guardar tu respuesta.")

    survey["sent_at_display"] = _fmt_dt(survey.get("sent_at"))
    survey["completed_at_display"] = _fmt_dt(survey.get("completed_at"))

    return render_template(
        "cases/survey.html",
        title=f"Encuesta {survey['case_id']}",
        survey=survey,
        error_message=error_message,
        success_message=success_message,
    )


REPORT_DEFAULT_RANGE_DAYS = 30
REPORT_MAX_RANGE_DAYS = 180
REPORT_MAX_EXPORT_ROWS = 5000
REPORT_MAX_PDF_ROWS = 1000
REPORT_KIND_CHOICES = [
    {
        'key': 'summary',
        'label': 'Resumen operativo',
        'description': 'Totales, estado, prioridad, equipos y responsables.',
    },
    {
        'key': 'cases',
        'label': 'Detalle de casos',
        'description': 'Listado detallado de tickets con ANS y responsables.',
    },
    {
        'key': 'closed',
        'label': 'Casos cerrados',
        'description': 'Consulta cierres por rango con filtro opcional por ID.',
    },
    {
        'key': 'survey_summary',
        'label': 'Calidad y satisfacción',
        'description': 'Resumen de encuestas e incluye detalle en PDF y Excel.',
    },
    {
        'key': 'survey_detail',
        'label': 'Detalle de encuestas',
        'description': 'Listado completo de respuestas y motivos.',
    },
    {
        'key': 'case_lookup',
        'label': 'Buscar caso',
        'description': 'Trae si fue atendido, encuesta de calidad, trazabilidad y adjuntos.',
    },
]


def _normalize_report_kind(raw_value: str | None) -> str:
    raw = text_value(raw_value).strip().lower()
    allowed = {item['key'] for item in REPORT_KIND_CHOICES}
    return raw if raw in allowed else 'summary'


def _parse_report_date(raw_value: str | None):
    raw = (raw_value or '').strip()
    if not raw:
        return None
    try:
        return datetime.strptime(raw, '%Y-%m-%d').date()
    except ValueError:
        return None


def _report_redirect_args(kind: str, date_from: str, date_to: str, closed_case_id: str = '', case_id_query: str = '') -> dict:
    args = {
        'kind': kind,
        'date_from': date_from,
        'date_to': date_to,
    }
    if closed_case_id:
        args['closed_case_id'] = closed_case_id
    if case_id_query:
        args['case_id_query'] = case_id_query
    return args


def _report_kind_choices() -> list[dict]:
    return [dict(item) for item in REPORT_KIND_CHOICES]


def _build_report_filter(kind: str = 'summary') -> dict:
    today = datetime.now().date()
    default_to = today
    default_from = today - timedelta(days=REPORT_DEFAULT_RANGE_DAYS - 1)

    raw_from = (request.args.get('date_from') or '').strip()
    raw_to = (request.args.get('date_to') or '').strip()
    closed_case_id = (request.args.get('closed_case_id') or '').strip()
    case_id_query = (request.args.get('case_id_query') or '').strip()

    messages: list[str] = []
    start_date = _parse_report_date(raw_from)
    end_date = _parse_report_date(raw_to)

    if raw_from and not start_date:
        messages.append('La fecha inicial no es válida. Se aplicó el rango por defecto.')
    if raw_to and not end_date:
        messages.append('La fecha final no es válida. Se aplicó el rango por defecto.')

    start_date = start_date or default_from
    end_date = end_date or default_to

    if start_date > end_date:
        start_date, end_date = end_date, start_date
        messages.append('El rango estaba invertido; se ajustó automáticamente.')

    requested_days = (end_date - start_date).days + 1
    if requested_days > REPORT_MAX_RANGE_DAYS:
        start_date = end_date - timedelta(days=REPORT_MAX_RANGE_DAYS - 1)
        messages.append(
            f'El rango máximo permitido es de {REPORT_MAX_RANGE_DAYS} días para proteger el rendimiento. '
            f'Se ajustó automáticamente desde {start_date.isoformat()} hasta {end_date.isoformat()}.'
        )

    visibility_sql, visibility_params = _visibility_condition('c')
    filters: list[str] = []
    params: list = []
    if visibility_sql:
        filters.append(visibility_sql)
        params.extend(visibility_params)

    if kind != 'case_lookup':
        date_column = 'c.created_at'
        if kind == 'closed':
            date_column = 'ISNULL(c.closed_at, c.updated_at)'
            filters.append("LOWER(c.status) = 'cerrado'")

        start_dt = datetime.combine(start_date, datetime.min.time())
        end_exclusive = datetime.combine(end_date + timedelta(days=1), datetime.min.time())
        filters.append(f"{date_column} >= ?")
        params.append(start_dt)
        filters.append(f"{date_column} < ?")
        params.append(end_exclusive)

        if kind == 'closed' and closed_case_id:
            filters.append('c.id LIKE ?')
            params.append(f'%{closed_case_id}%')

    where_clause = 'WHERE ' + ' AND '.join(filters) if filters else ''
    return {
        'kind': kind,
        'date_from': start_date.isoformat(),
        'date_to': end_date.isoformat(),
        'closed_case_id': closed_case_id,
        'case_id_query': case_id_query,
        'messages': messages,
        'where_clause': where_clause,
        'params': tuple(params),
        'requested_days': (end_date - start_date).days + 1,
    }


def _report_totals(filter_ctx: dict) -> dict:
    return select_one(
        f"""
        SELECT
            COUNT(*) AS total_cases,
            SUM(CASE WHEN LOWER(c.status) IN ('pendiente', 'asignado', 'reabierto', 'en espera de usuario') THEN 1 ELSE 0 END) AS open_cases,
            SUM(CASE WHEN LOWER(c.status) = 'resuelto' THEN 1 ELSE 0 END) AS resolved_cases,
            SUM(CASE WHEN LOWER(c.status) = 'reabierto' THEN 1 ELSE 0 END) AS reopened_cases,
            SUM(CASE WHEN LOWER(c.status) = 'cerrado' THEN 1 ELSE 0 END) AS closed_cases,
            SUM(CASE WHEN UPPER(c.priority) IN ('ALTA', 'P1') THEN 1 ELSE 0 END) AS high_priority
        FROM dbo.cases c
        {filter_ctx['where_clause']}
        """,
        filter_ctx['params'],
    ) or {}


def _report_by_team(filter_ctx: dict) -> list[dict]:
    return select_all(
        f"""
        SELECT c.assigned_team, COUNT(*) AS total_cases,
               SUM(CASE WHEN LOWER(c.status) IN ('pendiente', 'asignado', 'reabierto', 'en espera de usuario') THEN 1 ELSE 0 END) AS open_cases,
               SUM(CASE WHEN LOWER(c.status) = 'resuelto' THEN 1 ELSE 0 END) AS resolved_cases,
               SUM(CASE WHEN LOWER(c.status) = 'cerrado' THEN 1 ELSE 0 END) AS closed_cases
        FROM dbo.cases c
        {filter_ctx['where_clause']}
        GROUP BY c.assigned_team
        ORDER BY total_cases DESC, c.assigned_team
        """,
        filter_ctx['params'],
    )


def _report_by_priority(filter_ctx: dict) -> list[dict]:
    return select_all(
        f"""
        SELECT c.priority, COUNT(*) AS total_cases
        FROM dbo.cases c
        {filter_ctx['where_clause']}
        GROUP BY c.priority
        ORDER BY CASE UPPER(c.priority)
            WHEN 'ALTA' THEN 1
            WHEN 'MEDIA' THEN 2
            WHEN 'BAJA' THEN 3
            WHEN 'P1' THEN 1
            WHEN 'P2' THEN 2
            WHEN 'P3' THEN 3
            WHEN 'P4' THEN 4
            ELSE 99 END, c.priority
        """,
        filter_ctx['params'],
    )


def _report_by_status(filter_ctx: dict) -> list[dict]:
    return select_all(
        f"""
        SELECT c.status, COUNT(*) AS total_cases
        FROM dbo.cases c
        {filter_ctx['where_clause']}
        GROUP BY c.status
        ORDER BY CASE LOWER(c.status)
            WHEN 'pendiente' THEN 1
            WHEN 'asignado' THEN 2
            WHEN 'en espera de usuario' THEN 3
            WHEN 'resuelto' THEN 4
            WHEN 'reabierto' THEN 5
            WHEN 'cerrado' THEN 6
            ELSE 99 END, c.status
        """,
        filter_ctx['params'],
    )


def _report_top_owners(filter_ctx: dict) -> list[dict]:
    return select_all(
        f"""
        SELECT TOP 10 ISNULL(u.display_name, c.assigned_to) AS owner_name, COUNT(*) AS total_cases
        FROM dbo.cases c
        LEFT JOIN dbo.users u ON u.id = c.assigned_to
        {filter_ctx['where_clause']}
        GROUP BY ISNULL(u.display_name, c.assigned_to)
        ORDER BY total_cases DESC, owner_name
        """,
        filter_ctx['params'],
    )


def _survey_report_where(filter_ctx: dict) -> tuple[str, tuple]:
    visibility_sql, visibility_params = _visibility_condition('c')
    filters: list[str] = []
    params: list = []
    if visibility_sql:
        filters.append(visibility_sql)
        params.extend(visibility_params)

    start_date = datetime.strptime(filter_ctx['date_from'], '%Y-%m-%d').date()
    end_date = datetime.strptime(filter_ctx['date_to'], '%Y-%m-%d').date()
    start_dt = datetime.combine(start_date, datetime.min.time())
    end_exclusive = datetime.combine(end_date + timedelta(days=1), datetime.min.time())

    activity_at = 'COALESCE(s.completed_at, s.sent_at, s.created_at)'
    filters.append(f"{activity_at} >= ?")
    params.append(start_dt)
    filters.append(f"{activity_at} < ?")
    params.append(end_exclusive)

    where_clause = 'WHERE ' + ' AND '.join(filters) if filters else ''
    return where_clause, tuple(params)


def _survey_report_totals(filter_ctx: dict) -> dict:
    where_clause, params = _survey_report_where(filter_ctx)
    return select_one(
        f"""
        SELECT
            COUNT(*) AS total_surveys,
            SUM(CASE WHEN s.sent_at IS NOT NULL THEN 1 ELSE 0 END) AS sent_surveys,
            SUM(CASE WHEN s.completed_at IS NOT NULL THEN 1 ELSE 0 END) AS completed_surveys,
            SUM(CASE WHEN s.completed_at IS NULL THEN 1 ELSE 0 END) AS pending_surveys,
            AVG(CASE WHEN s.rating IS NOT NULL THEN CAST(s.rating AS DECIMAL(10,2)) END) AS avg_rating,
            SUM(CASE WHEN s.rating IN (1,2) THEN 1 ELSE 0 END) AS low_ratings,
            SUM(CASE WHEN s.rating IN (4,5) THEN 1 ELSE 0 END) AS high_ratings
        FROM dbo.case_surveys s
        INNER JOIN dbo.cases c ON c.id = s.case_id
        {where_clause}
        """,
        params,
    ) or {}


def _survey_report_by_rating(filter_ctx: dict) -> list[dict]:
    where_clause, params = _survey_report_where(filter_ctx)
    return select_all(
        f"""
        SELECT s.rating, COUNT(*) AS total_surveys
        FROM dbo.case_surveys s
        INNER JOIN dbo.cases c ON c.id = s.case_id
        {where_clause} AND s.rating IS NOT NULL
        GROUP BY s.rating
        ORDER BY s.rating
        """,
        params,
    )


def _survey_report_by_team(filter_ctx: dict) -> list[dict]:
    where_clause, params = _survey_report_where(filter_ctx)
    return select_all(
        f"""
        SELECT c.assigned_team,
               COUNT(*) AS total_surveys,
               SUM(CASE WHEN s.completed_at IS NOT NULL THEN 1 ELSE 0 END) AS completed_surveys,
               AVG(CASE WHEN s.rating IS NOT NULL THEN CAST(s.rating AS DECIMAL(10,2)) END) AS avg_rating
        FROM dbo.case_surveys s
        INNER JOIN dbo.cases c ON c.id = s.case_id
        {where_clause}
        GROUP BY c.assigned_team
        ORDER BY total_surveys DESC, c.assigned_team
        """,
        params,
    )


def _survey_report_detail_count(filter_ctx: dict) -> int:
    where_clause, params = _survey_report_where(filter_ctx)
    row = select_one(
        f"""
        SELECT COUNT(*) AS total_rows
        FROM dbo.case_surveys s
        INNER JOIN dbo.cases c ON c.id = s.case_id
        {where_clause}
        """,
        params,
    ) or {}
    return int(row.get('total_rows') or 0)


def _survey_report_detail_rows(filter_ctx: dict, limit: int | None = None) -> list[dict]:
    where_clause, params = _survey_report_where(filter_ctx)
    top_sql = f"TOP {int(limit)} " if limit else ""
    rows = select_all(
        f"""
        SELECT {top_sql}
               s.id AS survey_id,
               s.case_id,
               c.subject,
               c.status,
               c.assigned_team,
               ISNULL(u.display_name, c.assigned_to) AS assigned_to_name,
               c.requester_name,
               s.recipient_email,
               s.sent_at,
               s.completed_at,
               s.rating,
               s.reason,
               s.delivery_error,
               s.created_at
        FROM dbo.case_surveys s
        INNER JOIN dbo.cases c ON c.id = s.case_id
        LEFT JOIN dbo.users u ON u.id = c.assigned_to
        {where_clause}
        ORDER BY COALESCE(s.completed_at, s.sent_at, s.created_at) DESC, s.id DESC
        """,
        params,
    )
    for row in rows:
        row['assigned_team_label'] = role_label(row.get('assigned_team') or '') if row.get('assigned_team') else ''
        row['sent_at_display'] = _fmt_dt(row.get('sent_at'))
        row['completed_at_display'] = _fmt_dt(row.get('completed_at'))
        row['activity_at_display'] = row['completed_at_display'] if row.get('completed_at') else row['sent_at_display']
        row['rating_display'] = row.get('rating') if row.get('rating') is not None else '-'
        row['survey_status'] = 'Respondida' if row.get('completed_at') else ('Enviada' if row.get('sent_at') else 'Pendiente')
    return rows


def _report_detail_count(filter_ctx: dict) -> int:
    row = select_one(
        f"SELECT COUNT(*) AS total_rows FROM dbo.cases c {filter_ctx['where_clause']}",
        filter_ctx['params'],
    ) or {}
    return int(row.get('total_rows') or 0)


def _report_detail_rows(filter_ctx: dict, limit: int | None = None) -> list[dict]:
    top_sql = f"TOP {int(limit)} " if limit else ""
    if filter_ctx['kind'] == 'closed':
        rows = select_all(
            f"""
            SELECT {top_sql}c.id, c.status, c.priority, c.assigned_team,
                   ISNULL(u.display_name, c.assigned_to) AS assigned_to_name,
                   c.requester_name, c.requester_email, c.closed_at, c.subject, c.created_at, c.updated_at
            FROM dbo.cases c
            LEFT JOIN dbo.users u ON u.id = c.assigned_to
            {filter_ctx['where_clause']}
            ORDER BY c.closed_at DESC, c.created_at DESC
            """,
            filter_ctx['params'],
        )
        for row in rows:
            row['assigned_team_label'] = role_label(row.get('assigned_team') or '') if row.get('assigned_team') else ''
            row['closed_at_display'] = _fmt_dt(row.get('closed_at'))
            row['created_at_display'] = _fmt_dt(row.get('created_at'))
        return rows

    rows = select_all(
        f"""
        SELECT {top_sql}c.id, c.status, c.priority, c.assigned_team,
               ISNULL(u.display_name, c.assigned_to) AS assigned_to_name,
               c.requester_name, c.requester_email, c.created_at, c.updated_at, c.subject,
               c.sla_response_min, c.sla_resolution_min, c.response_due_at, c.resolution_due_at,
               c.first_response_at, c.resolved_at, c.closed_at
        FROM dbo.cases c
        LEFT JOIN dbo.users u ON u.id = c.assigned_to
        {filter_ctx['where_clause']}
        ORDER BY c.created_at DESC
        """,
        filter_ctx['params'],
    )
    decorated = []
    for row in rows:
        row = _decorate_case(row)
        row['created_at_display'] = _fmt_dt(row.get('created_at'))
        row['first_response_at_display'] = _fmt_dt(row.get('first_response_at'))
        row['resolved_at_display'] = _fmt_dt(row.get('resolved_at'))
        row['closed_at_display'] = _fmt_dt(row.get('closed_at'))
        decorated.append(row)
    return decorated


def _case_lookup_rows(case_id_query: str) -> list[dict]:
    query = text_value(case_id_query).strip()
    if not query:
        return []

    visibility_sql, visibility_params = _visibility_condition('c')
    filters: list[str] = []
    params: list = []
    if visibility_sql:
        filters.append(visibility_sql)
        params.extend(visibility_params)
    filters.append('c.id LIKE ?')
    params.append(f'%{query}%')
    where_clause = 'WHERE ' + ' AND '.join(filters)

    rows = select_all(
        f"""
        SELECT c.id, c.subject, c.description, c.status, c.priority, c.assigned_team,
               c.assigned_to, ISNULL(u.display_name, c.assigned_to) AS assigned_to_name,
               c.requester_name, c.requester_email,
               c.created_at, c.updated_at, c.first_response_at, c.resolved_at, c.closed_at,
               c.sla_response_min, c.sla_resolution_min, c.response_due_at, c.resolution_due_at,
               (SELECT COUNT(*) FROM dbo.case_updates cu WHERE cu.case_id = c.id) AS updates_count,
               (SELECT COUNT(*) FROM dbo.case_updates cu WHERE cu.case_id = c.id AND cu.is_solution = 1) AS solution_updates_count,
               (SELECT COUNT(*) FROM dbo.case_attachments ca WHERE ca.case_id = c.id) AS attachments_count,
               (SELECT COUNT(*) FROM dbo.case_surveys s WHERE s.case_id = c.id) AS surveys_count,
               (SELECT TOP 1 s.sent_at FROM dbo.case_surveys s WHERE s.case_id = c.id ORDER BY s.created_at DESC, s.id DESC) AS latest_survey_sent_at,
               (SELECT TOP 1 s.completed_at FROM dbo.case_surveys s WHERE s.case_id = c.id ORDER BY s.created_at DESC, s.id DESC) AS latest_survey_completed_at,
               (SELECT TOP 1 s.rating FROM dbo.case_surveys s WHERE s.case_id = c.id ORDER BY s.created_at DESC, s.id DESC) AS latest_survey_rating,
               (SELECT TOP 1 s.reason FROM dbo.case_surveys s WHERE s.case_id = c.id ORDER BY s.created_at DESC, s.id DESC) AS latest_survey_reason,
               (SELECT TOP 1 cu.message FROM dbo.case_updates cu WHERE cu.case_id = c.id AND cu.is_solution = 1 ORDER BY cu.created_at DESC, cu.id DESC) AS latest_solution_message
        FROM dbo.cases c
        LEFT JOIN dbo.users u ON u.id = c.assigned_to
        {where_clause}
        ORDER BY c.created_at DESC
        """,
        tuple(params),
    )

    for row in rows:
        _decorate_case(row)
        row['assigned_team_label'] = role_label(row.get('assigned_team') or '') if row.get('assigned_team') else ''
        row['created_at_display'] = _fmt_dt(row.get('created_at'))
        row['updated_at_display'] = _fmt_dt(row.get('updated_at'))
        row['first_response_at_display'] = _fmt_dt(row.get('first_response_at'))
        row['resolved_at_display'] = _fmt_dt(row.get('resolved_at'))
        row['closed_at_display'] = _fmt_dt(row.get('closed_at'))
        row['latest_survey_sent_at_display'] = _fmt_dt(row.get('latest_survey_sent_at'))
        row['latest_survey_completed_at_display'] = _fmt_dt(row.get('latest_survey_completed_at'))
        row['was_attended'] = bool(row.get('first_response_at') or row.get('resolved_at') or row.get('closed_at') or int(row.get('updates_count') or 0) > 0)
        row['has_quality_survey'] = int(row.get('surveys_count') or 0) > 0
        if row.get('latest_survey_completed_at'):
            row['survey_status'] = 'Respondida'
        elif row.get('latest_survey_sent_at'):
            row['survey_status'] = 'Enviada'
        else:
            row['survey_status'] = 'Sin encuesta'
    return rows


def _case_lookup_related(case_id: str) -> dict:
    updates = select_all(
        """
        SELECT TOP 20 id, author_name, author_email, message, is_solution, created_at
        FROM dbo.case_updates
        WHERE case_id = ?
        ORDER BY created_at DESC, id DESC
        """,
        (case_id,),
    )
    for row in updates:
        row['created_at_display'] = _fmt_dt(row.get('created_at'))

    surveys = select_all(
        """
        SELECT TOP 10 id AS survey_id, recipient_email, sent_at, completed_at, rating, reason, delivery_error, created_at
        FROM dbo.case_surveys
        WHERE case_id = ?
        ORDER BY created_at DESC, id DESC
        """,
        (case_id,),
    )
    for row in surveys:
        row['sent_at_display'] = _fmt_dt(row.get('sent_at'))
        row['completed_at_display'] = _fmt_dt(row.get('completed_at'))
        row['survey_status'] = 'Respondida' if row.get('completed_at') else ('Enviada' if row.get('sent_at') else 'Pendiente')

    attachments = select_all(
        """
        SELECT TOP 20 filename, size_bytes, uploaded_at
        FROM dbo.case_attachments
        WHERE case_id = ?
        ORDER BY uploaded_at DESC, id DESC
        """,
        (case_id,),
    )
    for row in attachments:
        row['uploaded_at_display'] = _fmt_dt(row.get('uploaded_at'))
    return {
        'updates': updates,
        'surveys': surveys,
        'attachments': attachments,
    }


def _build_case_lookup_payload(filter_ctx: dict) -> dict:
    rows = _case_lookup_rows(filter_ctx.get('case_id_query') or '')
    selected_case = rows[0] if rows else None
    related = {'updates': [], 'surveys': [], 'attachments': []}
    if selected_case:
        related = _case_lookup_related(selected_case['id'])
    return {
        'rows': rows,
        'selected_case': selected_case,
        'related': related,
    }


def _autosize_worksheet(ws):
    for column_cells in ws.columns:
        values = [str(cell.value) if cell.value is not None else '' for cell in column_cells]
        max_len = max((len(value) for value in values), default=0)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 42)


def _build_excel_report(kind: str, filter_ctx: dict, payload: dict):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    summary_title = f"Rango {filter_ctx['date_from']} a {filter_ctx['date_to']}"

    if kind == 'summary':
        ws = wb.active
        ws.title = 'Resumen'
        ws.append(['Reporte', 'Resumen operativo'])
        ws.append(['Rango', summary_title])
        ws.append(['Alcance', ', '.join(_role_scope_labels())])
        ws.append([])
        ws.append(['Indicador', 'Valor'])
        totals = payload['totals']
        metrics = [
            ('Total casos', totals.get('total_cases') or 0),
            ('Activos', totals.get('open_cases') or 0),
            ('Resueltos', totals.get('resolved_cases') or 0),
            ('Reabiertos', totals.get('reopened_cases') or 0),
            ('Cerrados', totals.get('closed_cases') or 0),
            ('Prioridad alta', totals.get('high_priority') or 0),
        ]
        for metric in metrics:
            ws.append(list(metric))
        for cell in ws[5]:
            cell.font = Font(bold=True)
        _autosize_worksheet(ws)

        sheets = [
            ('Por equipo', ['Equipo', 'Total', 'Activos', 'Resueltos', 'Cerrados'], payload['by_team'], lambda row: [role_label(row.get('assigned_team') or '') if row.get('assigned_team') else 'Sin equipo', row.get('total_cases') or 0, row.get('open_cases') or 0, row.get('resolved_cases') or 0, row.get('closed_cases') or 0]),
            ('Por estado', ['Estado', 'Total'], payload['by_status'], lambda row: [row.get('status') or 'Sin estado', row.get('total_cases') or 0]),
            ('Por prioridad', ['Prioridad', 'Total'], payload['by_priority'], lambda row: [normalize_priority(row.get('priority') or 'MEDIA'), row.get('total_cases') or 0]),
            ('Top responsables', ['Responsable', 'Total'], payload['top_owners'], lambda row: [row.get('owner_name') or 'Sin asignación', row.get('total_cases') or 0]),
        ]
        for title, headers, rows, mapper in sheets:
            wsx = wb.create_sheet(title=title)
            wsx.append(headers)
            for cell in wsx[1]:
                cell.font = Font(bold=True)
            for row in rows:
                wsx.append(mapper(row))
            _autosize_worksheet(wsx)
    elif kind == 'survey_summary':
        ws = wb.active
        ws.title = 'Encuesta'
        ws.append(['Reporte', 'Calidad y satisfacción'])
        ws.append(['Rango', summary_title])
        ws.append(['Alcance', ', '.join(_role_scope_labels())])
        ws.append([])
        ws.append(['Indicador', 'Valor'])
        totals = payload['totals']
        metrics = [
            ('Encuestas visibles', totals.get('total_surveys') or 0),
            ('Enviadas', totals.get('sent_surveys') or 0),
            ('Respondidas', totals.get('completed_surveys') or 0),
            ('Pendientes', totals.get('pending_surveys') or 0),
            ('Promedio', float(totals.get('avg_rating') or 0)),
            ('Calificación baja (1-2)', totals.get('low_ratings') or 0),
            ('Calificación alta (4-5)', totals.get('high_ratings') or 0),
        ]
        for metric in metrics:
            ws.append(list(metric))
        for cell in ws[5]:
            cell.font = Font(bold=True)
        _autosize_worksheet(ws)

        rating_ws = wb.create_sheet(title='Por calificación')
        rating_ws.append(['Calificación', 'Total'])
        for cell in rating_ws[1]:
            cell.font = Font(bold=True)
        for row in payload['by_rating']:
            rating_ws.append([row.get('rating') or '', row.get('total_surveys') or 0])
        _autosize_worksheet(rating_ws)

        team_ws = wb.create_sheet(title='Encuesta por equipo')
        team_ws.append(['Equipo', 'Encuestas', 'Respondidas', 'Promedio'])
        for cell in team_ws[1]:
            cell.font = Font(bold=True)
        for row in payload['by_team']:
            team_ws.append([
                role_label(row.get('assigned_team') or '') if row.get('assigned_team') else 'Sin equipo',
                row.get('total_surveys') or 0,
                row.get('completed_surveys') or 0,
                float(row.get('avg_rating') or 0),
            ])
        _autosize_worksheet(team_ws)

        detail_ws = wb.create_sheet(title='Detalle encuestas')
        detail_ws.append(['Encuesta', 'Caso', 'Equipo', 'Asignado a', 'Solicitante', 'Correo', 'Enviada', 'Respondida', 'Calificación', 'Motivo', 'Estado'])
        for cell in detail_ws[1]:
            cell.font = Font(bold=True)
        for row in payload.get('detail_rows') or []:
            detail_ws.append([
                row.get('survey_id') or '',
                row.get('case_id') or '',
                row.get('assigned_team_label') or '',
                row.get('assigned_to_name') or '',
                row.get('requester_name') or '',
                row.get('recipient_email') or '',
                row.get('sent_at_display') or '',
                row.get('completed_at_display') or '',
                row.get('rating') or '',
                row.get('reason') or '',
                row.get('survey_status') or '',
            ])
        _autosize_worksheet(detail_ws)
    elif kind == 'case_lookup':
        ws = wb.active
        ws.title = 'Buscar caso'
        ws.append(['Reporte', 'Buscar caso'])
        ws.append(['Consulta', filter_ctx.get('case_id_query') or ''])
        ws.append(['Alcance', ', '.join(_role_scope_labels())])
        ws.append([])
        ws.append(['ID', 'Estado', 'Prioridad', 'Atendido', 'Encuesta', 'Calificación', 'Equipo', 'Asignado a', 'Solicitante', 'Creado', 'Cerrado', 'Asunto'])
        for cell in ws[5]:
            cell.font = Font(bold=True)
        for row in payload.get('rows') or []:
            ws.append([
                row.get('id') or '',
                row.get('status') or '',
                row.get('priority') or '',
                'Sí' if row.get('was_attended') else 'No',
                row.get('survey_status') or '',
                row.get('latest_survey_rating') or '',
                row.get('assigned_team_label') or '',
                row.get('assigned_to_name') or '',
                row.get('requester_name') or '',
                row.get('created_at_display') or '',
                row.get('closed_at_display') or '',
                row.get('subject') or '',
            ])
        _autosize_worksheet(ws)

        selected_case = payload.get('selected_case')
        if selected_case:
            detail_ws = wb.create_sheet(title='Caso seleccionado')
            detail_ws.append(['Campo', 'Valor'])
            for cell in detail_ws[1]:
                cell.font = Font(bold=True)
            items = [
                ('Caso', selected_case.get('id') or ''),
                ('Asunto', selected_case.get('subject') or ''),
                ('Estado', selected_case.get('status') or ''),
                ('Prioridad', selected_case.get('priority') or ''),
                ('Equipo', selected_case.get('assigned_team_label') or ''),
                ('Asignado a', selected_case.get('assigned_to_name') or ''),
                ('Solicitante', selected_case.get('requester_name') or ''),
                ('Correo solicitante', selected_case.get('requester_email') or ''),
                ('Atendido', 'Sí' if selected_case.get('was_attended') else 'No'),
                ('Encuesta', selected_case.get('survey_status') or ''),
                ('Calificación más reciente', selected_case.get('latest_survey_rating') or ''),
                ('Motivo encuesta', selected_case.get('latest_survey_reason') or ''),
                ('Adjuntos', selected_case.get('attachments_count') or 0),
                ('Actualizaciones', selected_case.get('updates_count') or 0),
                ('Soluciones registradas', selected_case.get('solution_updates_count') or 0),
                ('Creado', selected_case.get('created_at_display') or ''),
                ('Primera respuesta', selected_case.get('first_response_at_display') or ''),
                ('Resuelto', selected_case.get('resolved_at_display') or ''),
                ('Cerrado', selected_case.get('closed_at_display') or ''),
                ('Descripción', selected_case.get('description') or ''),
                ('Última solución', selected_case.get('latest_solution_message') or ''),
            ]
            for item in items:
                detail_ws.append(list(item))
            _autosize_worksheet(detail_ws)

            updates_ws = wb.create_sheet(title='Trazabilidad')
            updates_ws.append(['Fecha', 'Autor', 'Solución', 'Mensaje'])
            for cell in updates_ws[1]:
                cell.font = Font(bold=True)
            for row in payload.get('related', {}).get('updates', []):
                updates_ws.append([
                    row.get('created_at_display') or '',
                    row.get('author_name') or row.get('author_email') or '',
                    'Sí' if row.get('is_solution') else 'No',
                    row.get('message') or '',
                ])
            _autosize_worksheet(updates_ws)
    else:
        ws = wb.active
        if kind == 'closed':
            title = 'Cerrados'
            report_label = 'Casos cerrados'
        elif kind == 'survey_detail':
            title = 'Detalle encuesta'
            report_label = 'Detalle de encuestas'
        else:
            title = 'Casos'
            report_label = 'Detalle de casos'
        ws.title = title
        ws.append(['Reporte', report_label])
        ws.append(['Rango', summary_title])
        ws.append(['Alcance', ', '.join(_role_scope_labels())])
        if kind == 'closed' and filter_ctx.get('closed_case_id'):
            ws.append(['Filtro caso', filter_ctx['closed_case_id']])
        ws.append([])
        if kind == 'closed':
            headers = ['ID', 'Estado', 'Prioridad', 'Equipo', 'Asignado a', 'Solicitante', 'Correo', 'Cerrado', 'Creado', 'Asunto']
            ws.append(headers)
            for row in payload['rows']:
                ws.append([
                    row.get('id') or '',
                    row.get('status') or '',
                    normalize_priority(row.get('priority') or 'MEDIA'),
                    row.get('assigned_team_label') or '',
                    row.get('assigned_to_name') or '',
                    row.get('requester_name') or '',
                    row.get('requester_email') or '',
                    row.get('closed_at_display') or '',
                    row.get('created_at_display') or '',
                    row.get('subject') or '',
                ])
        elif kind == 'survey_detail':
            headers = ['Encuesta', 'Caso', 'Equipo', 'Asignado a', 'Solicitante', 'Correo encuesta', 'Enviada', 'Respondida', 'Calificación', 'Motivo', 'Estado caso', 'Asunto']
            ws.append(headers)
            for row in payload['rows']:
                ws.append([
                    row.get('survey_id') or '',
                    row.get('case_id') or '',
                    row.get('assigned_team_label') or '',
                    row.get('assigned_to_name') or '',
                    row.get('requester_name') or '',
                    row.get('recipient_email') or '',
                    row.get('sent_at_display') or '',
                    row.get('completed_at_display') or '',
                    row.get('rating') or '',
                    row.get('reason') or '',
                    row.get('status') or '',
                    row.get('subject') or '',
                ])
        else:
            headers = ['ID', 'Estado', 'Prioridad', 'ANS respuesta', 'ANS resolución', 'Equipo', 'Asignado a', 'Solicitante', 'Correo', 'Creado', 'Primera respuesta', 'Resuelto', 'Asunto']
            ws.append(headers)
            for row in payload['rows']:
                ws.append([
                    row.get('id') or '',
                    row.get('status') or '',
                    row.get('priority') or '',
                    row.get('sla_response_text') or '',
                    row.get('sla_resolution_text') or '',
                    role_label(row.get('assigned_team') or '') if row.get('assigned_team') else '',
                    row.get('assigned_to_name') or '',
                    row.get('requester_name') or '',
                    row.get('requester_email') or '',
                    _fmt_dt(row.get('created_at')),
                    _fmt_dt(row.get('first_response_at')),
                    _fmt_dt(row.get('resolved_at')),
                    row.get('subject') or '',
                ])
        for cell in ws[5]:
            cell.font = Font(bold=True)
        _autosize_worksheet(ws)

    return wb


def _pdf_table(rows):
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle

    table = Table(rows, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F172A')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#CBD5E1')),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]
        )
    )
    return table


def _build_pdf_report(kind: str, filter_ctx: dict, payload: dict):
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    story = []

    titles = {
        'summary': 'Resumen operativo',
        'cases': 'Detalle de casos',
        'closed': 'Casos cerrados',
        'survey_summary': 'Calidad y satisfacción',
        'survey_detail': 'Detalle de encuestas',
        'case_lookup': 'Buscar caso',
    }
    story.append(Paragraph(titles[kind], styles['Title']))
    if kind == 'case_lookup':
        story.append(Paragraph(f"Consulta: {filter_ctx.get('case_id_query') or '-'}", styles['Normal']))
    else:
        story.append(Paragraph(f"Rango: {filter_ctx['date_from']} a {filter_ctx['date_to']}", styles['Normal']))
    story.append(Paragraph(f"Alcance: {', '.join(_role_scope_labels())}", styles['Normal']))
    story.append(Spacer(1, 12))

    if kind == 'summary':
        totals = payload['totals']
        story.append(_pdf_table([
            ['Indicador', 'Valor'],
            ['Total casos', totals.get('total_cases') or 0],
            ['Activos', totals.get('open_cases') or 0],
            ['Resueltos', totals.get('resolved_cases') or 0],
            ['Reabiertos', totals.get('reopened_cases') or 0],
            ['Cerrados', totals.get('closed_cases') or 0],
            ['Prioridad alta', totals.get('high_priority') or 0],
        ]))
        story.append(Spacer(1, 12))
        if payload['by_team']:
            story.append(Paragraph('Distribución por equipo', styles['Heading2']))
            story.append(_pdf_table([
                ['Equipo', 'Total', 'Activos', 'Resueltos', 'Cerrados'],
                *[[role_label(row.get('assigned_team') or '') if row.get('assigned_team') else 'Sin equipo', row.get('total_cases') or 0, row.get('open_cases') or 0, row.get('resolved_cases') or 0, row.get('closed_cases') or 0] for row in payload['by_team']]
            ]))
            story.append(Spacer(1, 12))
        if payload['by_status']:
            story.append(Paragraph('Distribución por estado', styles['Heading2']))
            story.append(_pdf_table([
                ['Estado', 'Total'],
                *[[row.get('status') or 'Sin estado', row.get('total_cases') or 0] for row in payload['by_status']]
            ]))
            story.append(Spacer(1, 12))
        if payload['by_priority']:
            story.append(Paragraph('Distribución por prioridad', styles['Heading2']))
            story.append(_pdf_table([
                ['Prioridad', 'Total'],
                *[[normalize_priority(row.get('priority') or 'MEDIA'), row.get('total_cases') or 0] for row in payload['by_priority']]
            ]))
            story.append(Spacer(1, 12))
        if payload['top_owners']:
            story.append(Paragraph('Responsables con mayor carga', styles['Heading2']))
            story.append(_pdf_table([
                ['Responsable', 'Total'],
                *[[row.get('owner_name') or 'Sin asignación', row.get('total_cases') or 0] for row in payload['top_owners']]
            ]))
    elif kind == 'survey_summary':
        totals = payload['totals']
        story.append(_pdf_table([
            ['Indicador', 'Valor'],
            ['Encuestas visibles', totals.get('total_surveys') or 0],
            ['Enviadas', totals.get('sent_surveys') or 0],
            ['Respondidas', totals.get('completed_surveys') or 0],
            ['Pendientes', totals.get('pending_surveys') or 0],
            ['Promedio', f"{float(totals.get('avg_rating') or 0):.2f}"],
            ['Calificación baja (1-2)', totals.get('low_ratings') or 0],
            ['Calificación alta (4-5)', totals.get('high_ratings') or 0],
        ]))
        story.append(Spacer(1, 12))
        if payload['by_rating']:
            story.append(Paragraph('Distribución por calificación', styles['Heading2']))
            story.append(_pdf_table([
                ['Calificación', 'Total'],
                *[[row.get('rating') or '', row.get('total_surveys') or 0] for row in payload['by_rating']]
            ]))
            story.append(Spacer(1, 12))
        if payload['by_team']:
            story.append(Paragraph('Encuesta por equipo', styles['Heading2']))
            story.append(_pdf_table([
                ['Equipo', 'Encuestas', 'Respondidas', 'Promedio'],
                *[[role_label(row.get('assigned_team') or '') if row.get('assigned_team') else 'Sin equipo', row.get('total_surveys') or 0, row.get('completed_surveys') or 0, f"{float(row.get('avg_rating') or 0):.2f}"] for row in payload['by_team']]
            ]))
            story.append(Spacer(1, 12))
        if payload.get('detail_rows'):
            story.append(Paragraph('Detalle de encuestas', styles['Heading2']))
            story.append(_pdf_table([
                ['Caso', 'Equipo', 'Solicitante', 'Correo', 'Enviada', 'Respondida', 'Calificación', 'Motivo'],
                *[[row.get('case_id') or '', row.get('assigned_team_label') or '', row.get('requester_name') or '', row.get('recipient_email') or '', row.get('sent_at_display') or '', row.get('completed_at_display') or '', row.get('rating') or '', row.get('reason') or ''] for row in payload['detail_rows']]
            ]))
    elif kind == 'survey_detail':
        story.append(Paragraph(f"Total filas: {len(payload['rows'])}", styles['Normal']))
        story.append(Spacer(1, 8))
        story.append(_pdf_table([
            ['Encuesta', 'Caso', 'Equipo', 'Asignado a', 'Solicitante', 'Correo', 'Enviada', 'Respondida', 'Calificación', 'Motivo'],
            *[[row.get('survey_id') or '', row.get('case_id') or '', row.get('assigned_team_label') or '', row.get('assigned_to_name') or '', row.get('requester_name') or '', row.get('recipient_email') or '', row.get('sent_at_display') or '', row.get('completed_at_display') or '', row.get('rating') or '', row.get('reason') or ''] for row in payload['rows']]
        ]))
    elif kind == 'closed':
        story.append(Paragraph(f"Total filas: {len(payload['rows'])}", styles['Normal']))
        story.append(Spacer(1, 8))
        story.append(_pdf_table([
            ['ID', 'Estado', 'Prioridad', 'Equipo', 'Asignado a', 'Solicitante', 'Correo', 'Cerrado', 'Asunto'],
            *[[row.get('id') or '', row.get('status') or '', normalize_priority(row.get('priority') or 'MEDIA'), row.get('assigned_team_label') or '', row.get('assigned_to_name') or '', row.get('requester_name') or '', row.get('requester_email') or '', row.get('closed_at_display') or '', row.get('subject') or ''] for row in payload['rows']]
        ]))
    elif kind == 'case_lookup':
        selected_case = payload.get('selected_case')
        rows = payload.get('rows') or []
        story.append(Paragraph(f"Coincidencias: {len(rows)}", styles['Normal']))
        story.append(Spacer(1, 8))
        story.append(_pdf_table([
            ['Caso', 'Estado', 'Prioridad', 'Atendido', 'Encuesta', 'Calificación', 'Equipo', 'Asignado a', 'Cerrado'],
            *[[row.get('id') or '', row.get('status') or '', row.get('priority') or '', 'Sí' if row.get('was_attended') else 'No', row.get('survey_status') or '', row.get('latest_survey_rating') or '', row.get('assigned_team_label') or '', row.get('assigned_to_name') or '', row.get('closed_at_display') or ''] for row in rows]
        ]))
        if selected_case:
            story.append(Spacer(1, 12))
            story.append(Paragraph('Detalle del caso seleccionado', styles['Heading2']))
            story.append(_pdf_table([
                ['Campo', 'Valor'],
                ['Caso', selected_case.get('id') or ''],
                ['Asunto', selected_case.get('subject') or ''],
                ['Estado', selected_case.get('status') or ''],
                ['Prioridad', selected_case.get('priority') or ''],
                ['Equipo', selected_case.get('assigned_team_label') or ''],
                ['Asignado a', selected_case.get('assigned_to_name') or ''],
                ['Solicitante', selected_case.get('requester_name') or ''],
                ['Correo solicitante', selected_case.get('requester_email') or ''],
                ['Atendido', 'Sí' if selected_case.get('was_attended') else 'No'],
                ['Encuesta', selected_case.get('survey_status') or ''],
                ['Calificación más reciente', selected_case.get('latest_survey_rating') or ''],
                ['Motivo encuesta', selected_case.get('latest_survey_reason') or ''],
                ['Primera respuesta', selected_case.get('first_response_at_display') or ''],
                ['Resuelto', selected_case.get('resolved_at_display') or ''],
                ['Cerrado', selected_case.get('closed_at_display') or ''],
                ['Adjuntos', selected_case.get('attachments_count') or 0],
                ['Actualizaciones', selected_case.get('updates_count') or 0],
            ]))
            related = payload.get('related', {})
            if related.get('surveys'):
                story.append(Spacer(1, 12))
                story.append(Paragraph('Encuestas de calidad', styles['Heading2']))
                story.append(_pdf_table([
                    ['Encuesta', 'Correo', 'Enviada', 'Respondida', 'Calificación', 'Motivo'],
                    *[[row.get('survey_id') or '', row.get('recipient_email') or '', row.get('sent_at_display') or '', row.get('completed_at_display') or '', row.get('rating') or '', row.get('reason') or ''] for row in related['surveys']]
                ]))
            if related.get('updates'):
                story.append(Spacer(1, 12))
                story.append(Paragraph('Trazabilidad reciente', styles['Heading2']))
                story.append(_pdf_table([
                    ['Fecha', 'Autor', 'Solución', 'Mensaje'],
                    *[[row.get('created_at_display') or '', row.get('author_name') or row.get('author_email') or '', 'Sí' if row.get('is_solution') else 'No', row.get('message') or ''] for row in related['updates']]
                ]))
    else:
        story.append(Paragraph(f"Total filas: {len(payload['rows'])}", styles['Normal']))
        story.append(Spacer(1, 8))
        story.append(_pdf_table([
            ['ID', 'Estado', 'Prioridad', 'ANS resp.', 'ANS resol.', 'Equipo', 'Asignado a', 'Solicitante', 'Creado', 'Primera respuesta', 'Resuelto', 'Asunto'],
            *[[row.get('id') or '', row.get('status') or '', row.get('priority') or '', row.get('sla_response_text') or '', row.get('sla_resolution_text') or '', role_label(row.get('assigned_team') or '') if row.get('assigned_team') else '', row.get('assigned_to_name') or '', row.get('requester_email') or row.get('requester_name') or '', _fmt_dt(row.get('created_at')), _fmt_dt(row.get('first_response_at')), _fmt_dt(row.get('resolved_at')), row.get('subject') or ''] for row in payload['rows']]
        ]))

    doc.build(story)
    buffer.seek(0)
    return buffer


def _build_selected_report(selected_kind: str, filter_ctx: dict) -> dict:
    if selected_kind == 'summary':
        return {
            'totals': _report_totals(filter_ctx),
            'by_team': _report_by_team(filter_ctx),
            'by_status': _report_by_status(filter_ctx),
            'by_priority': _report_by_priority(filter_ctx),
            'top_owners': _report_top_owners(filter_ctx),
        }
    if selected_kind == 'survey_summary':
        return {
            'totals': _survey_report_totals(filter_ctx),
            'by_rating': _survey_report_by_rating(filter_ctx),
            'by_team': _survey_report_by_team(filter_ctx),
            'detail_rows': _survey_report_detail_rows(filter_ctx, limit=200),
        }
    if selected_kind == 'survey_detail':
        return {
            'rows': _survey_report_detail_rows(filter_ctx, limit=200),
            'total_rows': _survey_report_detail_count(filter_ctx),
        }
    if selected_kind == 'closed':
        return {
            'rows': _report_detail_rows(filter_ctx, limit=200),
            'total_rows': _report_detail_count(filter_ctx),
        }
    if selected_kind == 'case_lookup':
        return _build_case_lookup_payload(filter_ctx)
    return {
        'rows': _report_detail_rows(filter_ctx, limit=200),
        'total_rows': _report_detail_count(filter_ctx),
    }


@cases_bp.route('/reports')
@login_required
def reports():
    selected_kind = _normalize_report_kind(request.args.get('kind') or 'summary')
    filter_kind = 'closed' if selected_kind == 'closed' else selected_kind
    if filter_kind not in {'summary', 'cases', 'closed', 'survey_summary', 'survey_detail', 'case_lookup'}:
        filter_kind = 'summary'
    filter_ctx = _build_report_filter(filter_kind)
    for message in filter_ctx['messages']:
        flash(message, 'warning')

    selected_payload = _build_selected_report(selected_kind, filter_ctx)
    selected_label = next((item['label'] for item in REPORT_KIND_CHOICES if item['key'] == selected_kind), 'Reporte')

    return render_template(
        'reports/index.html',
        title='Reportes',
        notif_count=_notif_count(),
        scope_labels=_role_scope_labels(),
        report_kinds=_report_kind_choices(),
        selected_kind=selected_kind,
        selected_label=selected_label,
        selected_payload=selected_payload,
        date_from=filter_ctx['date_from'],
        date_to=filter_ctx['date_to'],
        closed_case_id=filter_ctx['closed_case_id'],
        case_id_query=filter_ctx.get('case_id_query') or '',
        report_limits={
            'default_days': REPORT_DEFAULT_RANGE_DAYS,
            'max_days': REPORT_MAX_RANGE_DAYS,
            'max_export_rows': REPORT_MAX_EXPORT_ROWS,
            'max_pdf_rows': REPORT_MAX_PDF_ROWS,
        },
    )


@cases_bp.route('/reports/export')
@login_required
def reports_export():
    kind = _normalize_report_kind(request.args.get('kind') or 'summary')
    output_format = (request.args.get('format') or 'xlsx').strip().lower()
    if kind not in {'summary', 'cases', 'closed', 'survey_summary', 'survey_detail', 'case_lookup'}:
        flash('Tipo de reporte no válido.', 'warning')
        return redirect(url_for('cases.reports'))
    if output_format not in {'xlsx', 'pdf'}:
        flash('Formato de exportación no válido.', 'warning')
        return redirect(url_for('cases.reports'))

    filter_ctx = _build_report_filter('closed' if kind == 'closed' else kind)

    if kind == 'summary':
        payload = {
            'totals': _report_totals(filter_ctx),
            'by_team': _report_by_team(filter_ctx),
            'by_status': _report_by_status(filter_ctx),
            'by_priority': _report_by_priority(filter_ctx),
            'top_owners': _report_top_owners(filter_ctx),
        }
    elif kind == 'survey_summary':
        detail_limit = REPORT_MAX_PDF_ROWS if output_format == 'pdf' else REPORT_MAX_EXPORT_ROWS
        payload = {
            'totals': _survey_report_totals(filter_ctx),
            'by_rating': _survey_report_by_rating(filter_ctx),
            'by_team': _survey_report_by_team(filter_ctx),
            'detail_rows': _survey_report_detail_rows(filter_ctx, limit=detail_limit),
        }
    elif kind == 'case_lookup':
        payload = _build_case_lookup_payload(filter_ctx)
        if not (filter_ctx.get('case_id_query') or '').strip():
            flash('Debes indicar un caso para el reporte de búsqueda.', 'warning')
            return redirect(url_for('cases.reports', kind=kind, date_from=filter_ctx['date_from'], date_to=filter_ctx['date_to']))
    else:
        if kind == 'survey_detail':
            total_rows = _survey_report_detail_count(filter_ctx)
        else:
            total_rows = _report_detail_count(filter_ctx)
        row_limit = REPORT_MAX_PDF_ROWS if output_format == 'pdf' else REPORT_MAX_EXPORT_ROWS
        if total_rows > row_limit:
            flash(
                f'El reporte solicitado contiene {total_rows} filas. Reduce el rango para no exceder el límite de {row_limit} filas en {output_format.upper()}.',
                'warning',
            )
            return redirect(url_for('cases.reports', **_report_redirect_args(kind, filter_ctx['date_from'], filter_ctx['date_to'], filter_ctx.get('closed_case_id') or '', filter_ctx.get('case_id_query') or '')))
        if kind == 'survey_detail':
            payload = {'rows': _survey_report_detail_rows(filter_ctx)}
        else:
            payload = {'rows': _report_detail_rows(filter_ctx)}

    stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    safe_kind = {
        'summary': 'resumen',
        'cases': 'casos',
        'closed': 'cerrados',
        'survey_summary': 'encuesta_resumen',
        'survey_detail': 'encuesta_detalle',
        'case_lookup': 'buscar_caso',
    }[kind]

    if output_format == 'xlsx':
        wb = _build_excel_report(kind, filter_ctx, payload)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'reporte_{safe_kind}_{stamp}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    buffer = _build_pdf_report(kind, filter_ctx, payload)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'reporte_{safe_kind}_{stamp}.pdf',
        mimetype='application/pdf',
    )
